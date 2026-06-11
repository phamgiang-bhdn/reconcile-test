"""Excel export — tầng trình bày. Nhận rows/kpi (output của domain.logic), không đối soát lại.

Tiền giữ là ô SỐ nguyên (Excel SUM được), không phải text (E1). Dấu âm giữ nguyên (E4).
"""

import io
import re

from openpyxl import Workbook

_FORBIDDEN = re.compile(r"[*?:\\/\[\]]")

# cột tiền trong sheet ChiTiet (ghi dạng numeric)
_MONEY_COLS = ("gross_revenue", "refund_amount", "fee_total", "net_received")
_COLUMNS = (
    ("order_code", "Mã đơn"),
    ("order_status", "Trạng thái đơn"),
    ("settlement_date", "Ngày thanh toán"),
    ("gross_revenue", "Gross"),
    ("refund_amount", "Hoàn"),
    ("fee_total", "Phí"),
    ("net_received", "Net"),
    ("reconcile_status", "Đối soát"),
)
_KPI_LABELS = (
    ("total_gross", "Doanh thu gộp"),
    ("total_net", "Thực nhận (net)"),
    ("total_fees", "Tổng phí sàn"),
    ("reconciliation_rate", "Tỉ lệ đối soát"),
    ("refund_count", "Số đơn hoàn"),
    ("refund_total", "Tổng hoàn"),
)


def sanitize_sheet_name(name: str, default: str = "Sheet") -> str:
    """Loại ký tự Excel cấm (* ? : \\ / [ ]) và cắt <= 31 ký tự (E2)."""
    cleaned = _FORBIDDEN.sub("", name).strip()
    cleaned = cleaned[:31]
    return cleaned or default


def build_workbook_bytes(rows: list[dict], kpi: dict) -> bytes:
    wb = Workbook()

    ws_kpi = wb.active
    ws_kpi.title = sanitize_sheet_name("KPI")
    ws_kpi.append(["Chỉ số", "Giá trị (KPI toàn kỳ)"])
    for key, label in _KPI_LABELS:
        ws_kpi.append([label, kpi.get(key)])
        if key == "reconciliation_rate":
            # rate là tỉ lệ, hiển thị % cho khớp dashboard (giữ ô numeric)
            ws_kpi.cell(row=ws_kpi.max_row, column=2).number_format = "0.00%"

    ws = wb.create_sheet(sanitize_sheet_name("ChiTiet"))
    ws.append([label for _, label in _COLUMNS])
    for r in rows:
        ws.append([r.get(key) for key, _ in _COLUMNS])  # ô tiền là int → numeric (E1/E4)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
