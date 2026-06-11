import io

from openpyxl import load_workbook

from app.exports.reconciliation_export import build_workbook_bytes, sanitize_sheet_name


def test_sanitize_sheet_name():
    assert sanitize_sheet_name("Danh/Sách:Đơn*?[x]") == "DanhSáchĐơnx"
    assert sanitize_sheet_name("x" * 50) == "x" * 31
    assert sanitize_sheet_name("///") == "Sheet"  # rỗng sau sanitize -> default


def _sample():
    rows = [
        {"order_code": "A", "order_status": "completed", "settlement_date": "2026-06-01",
         "gross_revenue": 91000, "refund_amount": 0, "fee_total": -17560,
         "net_received": 73440, "reconcile_status": "matched"},
        {"order_code": "B", "order_status": "completed", "settlement_date": "2026-06-04",
         "gross_revenue": 101000, "refund_amount": -101000, "fee_total": -40000,
         "net_received": -40000, "reconcile_status": "refunded"},
        # dòng unsettled: mọi cột tiền None (như output reconcile() thật)
        {"order_code": "C", "order_status": "cancelled", "settlement_date": None,
         "gross_revenue": None, "refund_amount": None, "fee_total": None,
         "net_received": None, "reconcile_status": "unsettled"},
    ]
    kpi = {"total_gross": 192000, "total_net": 33440, "total_fees": -57560,
           "reconciliation_rate": 1.0, "refund_count": 1, "refund_total": -101000}
    return rows, kpi


def test_workbook_structure_and_numeric_money():
    rows, kpi = _sample()
    wb = load_workbook(io.BytesIO(build_workbook_bytes(rows, kpi)))
    assert wb.sheetnames == ["KPI", "ChiTiet"]

    ws = wb["ChiTiet"]
    assert ws.max_row == 1 + len(rows)  # header + rows

    # net của dòng refunded là ô SỐ âm (không phải text) -> Excel SUM được (E1/E4)
    net_b = ws.cell(row=3, column=7).value
    assert net_b == -40000
    assert isinstance(net_b, int)

    assert wb["KPI"].cell(row=3, column=2).value == 33440  # total_net

    # dòng unsettled (C): ô tiền None -> ô trống, KHÔNG vỡ workbook/SUM
    assert ws.cell(row=4, column=7).value is None


def test_empty_rows_still_valid():
    wb = load_workbook(io.BytesIO(build_workbook_bytes([], _sample()[1])))
    assert wb["ChiTiet"].max_row == 1  # chỉ header (E3)
